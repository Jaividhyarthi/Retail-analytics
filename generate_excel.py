"""
generate_excel.py
═══════════════════════════════════════════════════════
Retail Analytics — Excel Dashboard Generator
Run: python generate_excel.py
Output: Retail_Analytics_Dashboard.xlsx
═══════════════════════════════════════════════════════
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
import os, sys

# ── PATHS ────────────────────────────────────────────────────────
DATA_PATHS = [
    "data/Retail_Data_Transactions.csv",
    "exports/cleaned_transactions.csv",
    "Retail_Data_Transactions.csv",
]
OUTPUT_FILE = "Retail_Analytics_Dashboard.xlsx"

# ── COLORS ──────────────────────────────────────────────────────
NAVY     = "0D1B3E"
NAVY2    = "1A3A6B"
BLUE     = "1E6FA6"
ACCENT   = "42A5F5"
WHITE    = "FFFFFF"
OFFWHITE = "F0F4F8"
GRAY_LT  = "E2E8F0"
GREEN    = "22C55E"
ORANGE   = "F59E0B"
TEAL     = "0D9488"

# ── STYLE HELPERS ────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="0D1B3E", size=11, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def border_thin():
    s = Side(style="thin", color=GRAY_LT)
    return Border(left=s, right=s, top=s, bottom=s)

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def set_row_height(ws, row, height):
    ws.row_dimensions[row].height = height

def header_row(ws, row, cols_values, bg=NAVY, fg=WHITE, size=11):
    for col, val in enumerate(cols_values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill(bg)
        c.font = font(bold=True, color=fg, size=size)
        c.alignment = center()
        c.border = border_thin()

def data_row(ws, row, cols_values, alt=False):
    bg = OFFWHITE if alt else WHITE
    for col, val in enumerate(cols_values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill(bg)
        c.font = font(color=NAVY)
        c.alignment = center()
        c.border = border_thin()

# FIX: kpi_card now handles its own merging — no pre-merge needed before calling it.
# Each card occupies 2 rows × 2 columns. The value and label rows are merged
# separately so neither becomes a read-only MergedCell when written to.
def kpi_card(ws, row, col, value, label, bg=NAVY2, val_color=ACCENT):
    # Merge value row across 2 columns (e.g. A4:B4)
    ws.merge_cells(start_row=row,   start_column=col, end_row=row,   end_column=col+1)
    # Merge label row across 2 columns (e.g. A5:B5)
    ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)
    v = ws.cell(row=row,   column=col, value=value)
    v.fill = fill(bg); v.font = font(bold=True, color=val_color, size=20)
    v.alignment = center()
    l = ws.cell(row=row+1, column=col, value=label)
    l.fill = fill(bg); l.font = font(bold=False, color=WHITE, size=9)
    l.alignment = center()

# ════════════════════════════════════════════════════════════════
# LOAD & CLEAN DATA
# ════════════════════════════════════════════════════════════════
print("Loading data…")
df = None
for p in DATA_PATHS:
    if os.path.exists(p):
        df = pd.read_csv(p)
        print(f"  Loaded: {p}  ({len(df):,} rows)")
        break

if df is None:
    print("❌  No CSV found. Place Retail_Data_Transactions.csv in data/ folder.")
    sys.exit(1)

# Step 1: strip and lowercase ALL column names, remove every separator
df.columns = (df.columns
                .str.strip()
                .str.lower()
                .str.replace(r"[\s\-_/\\]", "", regex=True))

print(f"  Raw columns: {list(df.columns)}")

# Step 2: find each needed column by checking many possible names
def find_col(df, *candidates):
    for c in candidates:
        if c in df.columns:
            return c
    return None

tid_col   = find_col(df, "transactionid","txnid","orderid","id","transid")
time_col  = find_col(df, "transactiontime","transactiondate","transdate",
                        "datetime","date","timestamp","orderdate","saledate")
icode_col = find_col(df, "itemcode","sku","productcode","prodcode","code")
idesc_col = find_col(df, "itemdescription","description","productname",
                        "itemname","product","prodname","name")
qty_col   = find_col(df, "numberofitemspurchased","qty","quantity","units",
                        "count","numitems","noofitems","items")
price_col = find_col(df, "costperitem","unitprice","price","cost",
                        "amount","tranamount","value","unitcost","saleamount")
cty_col   = find_col(df, "country","region","market","location","territory")

print(f"  Detected → id:{tid_col} | time:{time_col} | qty:{qty_col} | price:{price_col} | country:{cty_col}")

# Step 3: rename detected columns to standard names
rmap = {}
if tid_col:   rmap[tid_col]   = "transaction_id"
if time_col:  rmap[time_col]  = "transaction_time"
if icode_col: rmap[icode_col] = "item_code"
if idesc_col: rmap[idesc_col] = "item_description"
if qty_col:   rmap[qty_col]   = "qty"
if price_col: rmap[price_col] = "cost_per_item"
if cty_col:   rmap[cty_col]   = "country"
df.rename(columns=rmap, inplace=True)

# Step 4: fill in any missing columns with defaults
if "transaction_id"   not in df.columns: df["transaction_id"]   = df.index.astype(str)
if "item_code"        not in df.columns: df["item_code"]        = "ITEM001"
if "item_description" not in df.columns: df["item_description"] = "Product"
if "country"          not in df.columns: df["country"]          = "Unknown"

# Step 5: fix qty and cost_per_item — handle missing cols safely
if "qty" not in df.columns:
    df["qty"] = 1
else:
    df["qty"] = pd.to_numeric(df["qty"], errors="coerce").fillna(1)

if "cost_per_item" not in df.columns:
    num_cols = [c for c in df.select_dtypes(include="number").columns
                if c not in ("qty",)]
    if num_cols:
        df["cost_per_item"] = pd.to_numeric(df[num_cols[0]], errors="coerce").fillna(0)
        print(f"  ⚠ Using '{num_cols[0]}' as cost_per_item fallback")
    else:
        df["cost_per_item"] = 1.0
else:
    df["cost_per_item"] = pd.to_numeric(df["cost_per_item"], errors="coerce").fillna(0)

# Step 6: parse datetime
if "transaction_time" in df.columns:
    df["transaction_time"] = pd.to_datetime(df["transaction_time"], errors="coerce")
    df.dropna(subset=["transaction_time"], inplace=True)
else:
    print("⚠  No date column found — using today as placeholder.")
    df["transaction_time"] = pd.Timestamp("today")

# Step 7: filter invalid rows
df = df[(df["qty"] > 0) & (df["cost_per_item"] > 0)]
print(f"  Clean rows: {len(df):,}")

# Step 8: calculated fields
df["total_sales"]  = df["qty"] * df["cost_per_item"]
df["txn_year"]     = df["transaction_time"].dt.year
df["txn_month"]    = df["transaction_time"].dt.month
df["txn_dow"]      = df["transaction_time"].dt.day_name()
df["year_month"]   = df["transaction_time"].dt.to_period("M").astype(str)

# ════════════════════════════════════════════════════════════════
# AGGREGATIONS
# ════════════════════════════════════════════════════════════════
total_rev   = df["total_sales"].sum()
total_txns  = df["transaction_id"].nunique()
n_countries = df["country"].nunique()
n_products  = df["item_code"].nunique()
aov         = df.groupby("transaction_id")["total_sales"].sum().mean()
avg_basket  = df.groupby("transaction_id")["qty"].sum().mean()

monthly = (df.groupby("year_month")
             .agg(Revenue=("total_sales","sum"),
                  Transactions=("transaction_id","count"))
             .reset_index()
             .sort_values("year_month"))

country = (df.groupby("country")
             .agg(Revenue=("total_sales","sum"),
                  Transactions=("transaction_id","count"),
                  Units=("qty","sum"),
                  AOV=("total_sales","mean"))
             .reset_index()
             .sort_values("Revenue", ascending=False))

product = (df.groupby(["item_code","item_description"])
             .agg(Revenue=("total_sales","sum"),
                  Units=("qty","sum"),
                  Transactions=("transaction_id","count"),
                  AvgPrice=("cost_per_item","mean"))
             .reset_index()
             .sort_values("Revenue", ascending=False))

dow_order = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
dow = (df.groupby("txn_dow")["total_sales"]
         .sum().reindex(dow_order)
         .reset_index()
         .dropna())

print("  Aggregations done.")

# ════════════════════════════════════════════════════════════════
# BUILD WORKBOOK
# ════════════════════════════════════════════════════════════════
wb = Workbook()

# ────────────────────────────────────────────────────────────────
# SHEET 1 — DASHBOARD
# ────────────────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Dashboard"
ws1.sheet_view.showGridLines = False

# Title
ws1.merge_cells("A1:L1")
t = ws1["A1"]
t.value = "RETAIL CHAIN SALES ANALYTICS DASHBOARD"
t.fill = fill(NAVY); t.font = font(bold=True, color=WHITE, size=18)
t.alignment = center()
set_row_height(ws1, 1, 48)

ws1.merge_cells("A2:L2")
s = ws1["A2"]
s.value = "InternshipStudio Final Project  ·  Python · SQL · Excel · Streamlit"
s.fill = fill(NAVY2); s.font = font(italic=True, color=ACCENT, size=11)
s.alignment = center()
set_row_height(ws1, 2, 24)

set_row_height(ws1, 3, 12)

# KPI cards — kpi_card handles all merging internally now
kpi_card(ws1, 4, 1,  f"${total_rev:,.0f}",  "Total Revenue",    NAVY,  ACCENT)
kpi_card(ws1, 4, 3,  f"{total_txns:,}",      "Transactions",     NAVY2, "90CAF9")
kpi_card(ws1, 4, 5,  f"{n_countries}",       "Countries",        BLUE,  WHITE)
kpi_card(ws1, 4, 7,  f"{n_products}",        "Products",         TEAL,  WHITE)
kpi_card(ws1, 4, 9,  f"${aov:,.2f}",         "Avg Order Value",  NAVY,  ORANGE)
kpi_card(ws1, 4, 11, f"{avg_basket:.1f}",    "Avg Basket Size",  NAVY2, GREEN)
for r in [4, 5]: set_row_height(ws1, r, 32)

set_row_height(ws1, 6, 10)

# Monthly revenue table
ws1.merge_cells("A7:C7")
h = ws1["A7"]
h.value = "Monthly Revenue"; h.fill = fill(NAVY2)
h.font = font(bold=True, color=WHITE, size=11); h.alignment = center()
set_row_height(ws1, 7, 22)

header_row(ws1, 8, ["Month", "Revenue ($)", "Transactions"])
for i, row in enumerate(monthly.itertuples(), 9):
    data_row(ws1, i, [row.year_month, round(row.Revenue, 2), row.Transactions], alt=i%2==0)

# Line chart
chart1 = LineChart()
chart1.title = "Monthly Revenue Trend"
chart1.style = 10
chart1.y_axis.title = "Revenue ($)"
chart1.x_axis.title = "Month"
chart1.width = 20; chart1.height = 13
data1  = Reference(ws1, min_col=2, min_row=8, max_row=8+len(monthly))
cats1  = Reference(ws1, min_col=1, min_row=9, max_row=8+len(monthly))
chart1.add_data(data1, titles_from_data=True)
chart1.set_categories(cats1)
chart1.series[0].graphicalProperties.line.solidFill = BLUE
chart1.series[0].graphicalProperties.line.width     = 25000
ws1.add_chart(chart1, "E7")

# Country revenue table
cstart = 9 + len(monthly) + 2
ws1.merge_cells(f"A{cstart}:E{cstart}")
h2 = ws1[f"A{cstart}"]
h2.value = "Revenue by Country"; h2.fill = fill(NAVY2)
h2.font  = font(bold=True, color=WHITE, size=11); h2.alignment = center()
set_row_height(ws1, cstart, 22)

header_row(ws1, cstart+1, ["Country","Revenue ($)","Transactions","Units","AOV ($)"])
for i, row in enumerate(country.itertuples(), cstart+2):
    data_row(ws1, i, [
        row.country, round(row.Revenue,2),
        row.Transactions, int(row.Units), round(row.AOV,2)
    ], alt=i%2==0)

# Bar chart — country
chart2 = BarChart()
chart2.type = "bar"
chart2.title = "Revenue by Country"
chart2.style = 10; chart2.width = 20; chart2.height = 13
data2  = Reference(ws1, min_col=2, min_row=cstart+1, max_row=cstart+1+len(country))
cats2  = Reference(ws1, min_col=1, min_row=cstart+2, max_row=cstart+1+len(country))
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats2)
chart2.series[0].graphicalProperties.solidFill = BLUE
ws1.add_chart(chart2, f"E{cstart}")

for col, w in [(1,14),(2,18),(3,16),(4,14),(5,14)]:
    set_col_width(ws1, col, w)

# ────────────────────────────────────────────────────────────────
# SHEET 2 — PRODUCTS
# ────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Products")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:G1")
ws2["A1"].value = "PRODUCT PERFORMANCE ANALYSIS"
ws2["A1"].fill  = fill(TEAL)
ws2["A1"].font  = font(bold=True, color=WHITE, size=16)
ws2["A1"].alignment = center()
set_row_height(ws2, 1, 40)

header_row(ws2, 2, ["Item Code","Description","Revenue ($)","Units","Transactions","Avg Price ($)","% of Total"])
set_row_height(ws2, 2, 22)

total_for_pct = product["Revenue"].sum()
for i, row in enumerate(product.head(30).itertuples(), 3):
    pct = row.Revenue / total_for_pct * 100
    data_row(ws2, i, [
        row.item_code, row.item_description,
        round(row.Revenue, 2), int(row.Units),
        row.Transactions, round(row.AvgPrice, 2),
        round(pct, 2)
    ], alt=i%2==0)

ws2.conditional_formatting.add(
    f"C3:C{2+min(30,len(product))}",
    DataBarRule(start_type="min", end_type="max", color=BLUE)
)

chart3 = BarChart()
chart3.type = "bar"
chart3.title = "Top 15 Products — Revenue"
chart3.style = 10; chart3.width = 24; chart3.height = 16
data3  = Reference(ws2, min_col=3, min_row=2, max_row=17)
cats3  = Reference(ws2, min_col=2, min_row=3, max_row=17)
chart3.add_data(data3, titles_from_data=True)
chart3.set_categories(cats3)
chart3.series[0].graphicalProperties.solidFill = TEAL
ws2.add_chart(chart3, "I2")

for col, w in [(1,15),(2,40),(3,16),(4,12),(5,14),(6,14),(7,12)]:
    set_col_width(ws2, col, w)

# ────────────────────────────────────────────────────────────────
# SHEET 3 — TIME ANALYSIS
# ────────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("Time Analysis")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:D1")
ws3["A1"].value = "TIME-SERIES ANALYSIS"
ws3["A1"].fill  = fill(NAVY)
ws3["A1"].font  = font(bold=True, color=WHITE, size=16)
ws3["A1"].alignment = center()
set_row_height(ws3, 1, 40)

header_row(ws3, 2, ["Year-Month","Revenue ($)","Transactions"])
for i, row in enumerate(monthly.itertuples(), 3):
    data_row(ws3, i, [row.year_month, round(row.Revenue,2), row.Transactions], alt=i%2==0)

chart4 = LineChart()
chart4.title = "Monthly Revenue Trend"
chart4.style = 10; chart4.width = 26; chart4.height = 15
data4  = Reference(ws3, min_col=2, min_row=2, max_row=2+len(monthly))
cats4  = Reference(ws3, min_col=1, min_row=3, max_row=2+len(monthly))
chart4.add_data(data4, titles_from_data=True)
chart4.set_categories(cats4)
chart4.series[0].graphicalProperties.line.solidFill = BLUE
chart4.series[0].graphicalProperties.line.width     = 28000
ws3.add_chart(chart4, "F1")

# Day of week
dstart = 3 + len(monthly) + 2
ws3.merge_cells(f"A{dstart}:D{dstart}")
ws3[f"A{dstart}"].value = "Day of Week Revenue"
ws3[f"A{dstart}"].fill  = fill(NAVY2)
ws3[f"A{dstart}"].font  = font(bold=True, color=WHITE)
ws3[f"A{dstart}"].alignment = center()
set_row_height(ws3, dstart, 22)

header_row(ws3, dstart+1, ["Day","Revenue ($)"])
for i, row in enumerate(dow.itertuples(), dstart+2):
    data_row(ws3, i, [row.txn_dow, round(row.total_sales, 2)], alt=i%2==0)

chart5 = BarChart()
chart5.title = "Revenue by Day of Week"
chart5.style = 10; chart5.width = 26; chart5.height = 13
data5  = Reference(ws3, min_col=2, min_row=dstart+1, max_row=dstart+8)
cats5  = Reference(ws3, min_col=1, min_row=dstart+2, max_row=dstart+8)
chart5.add_data(data5, titles_from_data=True)
chart5.set_categories(cats5)
chart5.series[0].graphicalProperties.solidFill = ORANGE
ws3.add_chart(chart5, f"F{dstart}")

for col, w in [(1,15),(2,18),(3,16)]:
    set_col_width(ws3, col, w)

# ────────────────────────────────────────────────────────────────
# SHEET 4 — GEOGRAPHIC
# ────────────────────────────────────────────────────────────────
ws4 = wb.create_sheet("Geographic")
ws4.sheet_view.showGridLines = False

ws4.merge_cells("A1:F1")
ws4["A1"].value = "GEOGRAPHIC REVENUE ANALYSIS"
ws4["A1"].fill  = fill(BLUE)
ws4["A1"].font  = font(bold=True, color=WHITE, size=16)
ws4["A1"].alignment = center()
set_row_height(ws4, 1, 40)

header_row(ws4, 2, ["Country","Revenue ($)","Transactions","Units","AOV ($)","Revenue %"])
total_rev_c = country["Revenue"].sum()
for i, row in enumerate(country.itertuples(), 3):
    pct = row.Revenue / total_rev_c * 100
    data_row(ws4, i, [
        row.country, round(row.Revenue,2),
        row.Transactions, int(row.Units),
        round(row.AOV,2), round(pct,1)
    ], alt=i%2==0)

ws4.conditional_formatting.add(
    f"F3:F{2+len(country)}",
    ColorScaleRule(start_type="min", start_color="FFFFFF",
                   end_type="max",   end_color=BLUE)
)

chart6 = BarChart()
chart6.type = "bar"
chart6.title = "Revenue by Country"
chart6.style = 10; chart6.width = 22; chart6.height = 14
data6  = Reference(ws4, min_col=2, min_row=2, max_row=2+len(country))
cats6  = Reference(ws4, min_col=1, min_row=3, max_row=2+len(country))
chart6.add_data(data6, titles_from_data=True)
chart6.set_categories(cats6)
chart6.series[0].graphicalProperties.solidFill = BLUE
ws4.add_chart(chart6, "H2")

chart7 = PieChart()
chart7.title = "Revenue Share by Country"
chart7.style = 10; chart7.width = 14; chart7.height = 12
data7  = Reference(ws4, min_col=2, min_row=2, max_row=2+len(country))
cats7  = Reference(ws4, min_col=1, min_row=3, max_row=2+len(country))
chart7.add_data(data7, titles_from_data=True)
chart7.set_categories(cats7)
ws4.add_chart(chart7, "H20")

for col, w in [(1,18),(2,18),(3,16),(4,14),(5,14),(6,14)]:
    set_col_width(ws4, col, w)

# ────────────────────────────────────────────────────────────────
# SHEET 5 — DATA SUMMARY
# ────────────────────────────────────────────────────────────────
ws5 = wb.create_sheet("Data Summary")
ws5.sheet_view.showGridLines = False

ws5.merge_cells("A1:C1")
ws5["A1"].value = "DATA QUALITY & SUMMARY REPORT"
ws5["A1"].fill  = fill(NAVY)
ws5["A1"].font  = font(bold=True, color=WHITE, size=14)
ws5["A1"].alignment = center()
set_row_height(ws5, 1, 36)

summary_rows = [
    ["Total Rows",             f"{len(df):,}"],
    ["Unique Transactions",    f"{total_txns:,}"],
    ["Unique Products",        f"{n_products}"],
    ["Countries",              f"{n_countries}"],
    ["Date Range",             f"{df['transaction_time'].min().date()} → {df['transaction_time'].max().date()}"],
    ["Min Price",              f"${df['cost_per_item'].min():.2f}"],
    ["Max Price",              f"${df['cost_per_item'].max():.2f}"],
    ["Avg Price",              f"${df['cost_per_item'].mean():.2f}"],
    ["Median Price",           f"${df['cost_per_item'].median():.2f}"],
    ["Total Revenue",          f"${total_rev:,.2f}"],
    ["Avg Order Value",        f"${aov:.2f}"],
    ["Avg Basket Size",        f"{avg_basket:.1f} items"],
    ["Cleaning Applied",       "Nulls dropped, invalid qty/price removed"],
    ["Calculated Fields",      "total_sales, year, month, DOW, hour, year_month"],
]

header_row(ws5, 2, ["Metric", "Value"])
for i, (k, v) in enumerate(summary_rows, 3):
    bg = OFFWHITE if i%2==0 else WHITE
    ws5.cell(row=i, column=1, value=k).fill  = fill(bg)
    ws5.cell(row=i, column=1).font           = font(bold=True, color=NAVY)
    ws5.cell(row=i, column=1).border         = border_thin()
    ws5.cell(row=i, column=2, value=v).fill  = fill(bg)
    ws5.cell(row=i, column=2).font           = font(color=BLUE)
    ws5.cell(row=i, column=2).border         = border_thin()
    ws5.cell(row=i, column=2).alignment      = center()

set_col_width(ws5, 1, 28)
set_col_width(ws5, 2, 40)

# ── SAVE ─────────────────────────────────────────────────────────
wb.save(OUTPUT_FILE)
print(f"\n✅  {OUTPUT_FILE} created!")
print(f"    Sheets: {[ws.title for ws in wb.worksheets]}")
print(f"\n📁  Add to Google Drive submission folder:")
print(f"    · {OUTPUT_FILE}")
print(f"    · Retail_Analytics_Presentation.pptx")
print(f"    · REPORT.md (convert to PDF)")
print(f"    · Streamlit live URL")